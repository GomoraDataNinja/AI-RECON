import io
import os
import re
import json
import hashlib
from dataclasses import dataclass
from datetime import datetime, date
from typing import Optional, List, Dict, Any, Tuple
from contextlib import contextmanager

import numpy as np
import pandas as pd
import streamlit as st
from dateutil.parser import parse as dt_parse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from copy import copy as pycopy

# Ollama (local)
try:
    import requests
except Exception:
    requests = None


# =========================
# Streamlit compatibility
# =========================
def safe_rerun():
    try:
        st.rerun()
    except Exception:
        try:
            st.experimental_rerun()
        except Exception:
            pass


def safe_chat_input(label: str, key: str = "chat_input_fallback"):
    try:
        return st.chat_input(label)
    except Exception:
        form_key = f"{key}_form"
        txt_key = f"{key}_txt"
        with st.form(form_key, clear_on_submit=True):
            txt = st.text_input(label, key=txt_key)
            submitted = st.form_submit_button("Send")
        if submitted:
            txt = (txt or "").strip()
            return txt if txt else None
        return None


@contextmanager
def safe_chat_message(role: str):
    try:
        with st.chat_message(role):
            yield
    except Exception:
        with st.container():
            st.markdown(f"**{role.capitalize()}**")
            yield


# =========================
# Config
# =========================
@dataclass(frozen=True)
class ReconConfig:
    date_window_days: int = 14
    amount_tolerance: float = 0.05
    min_auto_confidence: float = 0.70
    flip_ledger_sign: bool = True

    # perf guards
    max_combo_size: int = 4
    max_candidates_per_bucket: int = 20
    min_token_len: int = 4


# =========================
# Theme
# =========================
WF_RED = "#D71E28"
WF_BG = "#F3F3F3"
WF_CARD = "#FFFFFF"
WF_TEXT = "#111111"
WF_MUTED = "#666666"
WF_BORDER = "#E6E6E6"


# =========================
# Helpers
# =========================
def to_str(x) -> str:
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    return str(x).strip()


def norm_space(s: str) -> str:
    s = to_str(s)
    return re.sub(r"\s+", " ", s).strip()


def upper(s: str) -> str:
    return norm_space(s).upper()


def norm_ref(s: str) -> str:
    s = upper(s)
    s = re.sub(r"\s+", "", s)
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s


def to_num(x):
    if x is None:
        return np.nan
    try:
        if pd.isna(x):
            return np.nan
    except Exception:
        pass

    if isinstance(x, (int, float, np.number)) and not (isinstance(x, float) and np.isnan(x)):
        return float(x)

    s = to_str(x)
    if not s:
        return np.nan

    s = s.replace(",", "").replace(" ", "")
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]

    s = re.sub(r"[^\d\.\-]", "", s)
    if s in ("", "-", "."):
        return np.nan

    try:
        v = float(s)
        return -abs(v) if neg else v
    except Exception:
        return np.nan


def to_date(x, dayfirst=True):
    if isinstance(x, (pd.Timestamp, np.datetime64)):
        return pd.to_datetime(x, errors="coerce")
    s = to_str(x)
    if not s:
        return pd.NaT
    try:
        return pd.to_datetime(dt_parse(s, dayfirst=dayfirst, fuzzy=True))
    except Exception:
        return pd.NaT


def round2(x):
    try:
        if pd.isna(x):
            return np.nan
    except Exception:
        pass
    try:
        return float(round(float(x), 2))
    except Exception:
        return np.nan


def round0(x):
    try:
        if pd.isna(x):
            return np.nan
    except Exception:
        pass
    try:
        return float(round(float(x), 0))
    except Exception:
        return np.nan


def within_window(d1, d2, days: int) -> bool:
    if pd.isna(d1) or pd.isna(d2):
        return False
    return abs((pd.to_datetime(d1) - pd.to_datetime(d2)).days) <= int(days)


def hash_row(*parts) -> str:
    raw = "|".join([to_str(p) for p in parts])
    return hashlib.md5(raw.encode("utf-8")).hexdigest()[:16]


def norm_header_key(h: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", upper(h))


def safe_col(df: pd.DataFrame, col: str) -> pd.Series:
    if col and col in df.columns:
        return df[col]
    return pd.Series([None] * len(df))


def extract_tokens(text: str, min_len: int = 4) -> List[str]:
    t = upper(text)
    parts = re.findall(r"[A-Z0-9]+", t)
    return [p for p in parts if len(p) >= min_len]


def token_overlap(a: str, b: str, min_len: int = 4) -> float:
    ta = set(extract_tokens(a, min_len=min_len))
    tb = set(extract_tokens(b, min_len=min_len))
    if not ta or not tb:
        return 0.0
    return float(len(ta.intersection(tb)) / max(1, len(ta.union(tb))))


# =========================
# More DocID patterns (flexible)
# =========================
DOCID_PATTERNS = [
    re.compile(r"\b(HREINV|HRECRN)\s*0*([0-9]+)\b", re.IGNORECASE),
    re.compile(r"\b(INV|INVOICE)\s*[-_ ]*\s*0*([0-9]{4,})\b", re.IGNORECASE),
    re.compile(r"\b(CRN|CREDITNOTE|CREDIT)\s*[-_ ]*\s*0*([0-9]{4,})\b", re.IGNORECASE),
    re.compile(r"\b(DN|DEBITNOTE|DEBIT)\s*[-_ ]*\s*0*([0-9]{4,})\b", re.IGNORECASE),
    re.compile(r"\b(DOC|DOCUMENT)\s*(NO|NUMBER|#)?\s*[:\- ]*\s*0*([0-9]{4,})\b", re.IGNORECASE),
    re.compile(r"\bAE[A-Z]{0,3}\d{4,}\b", re.IGNORECASE),
    re.compile(r"\b(?:INV|DOC|REF|REFERENCE)\D{0,6}(0*[0-9]{6,})\b", re.IGNORECASE),
]


def extract_docid(text: str) -> str:
    t = upper(text)
    best = ""
    best_pos = -1

    for rx in DOCID_PATTERNS:
        for m in rx.finditer(t):
            pos = m.start()
            if rx.pattern.startswith("\\bAE"):
                cand = norm_ref(m.group(0))
            else:
                g = [x for x in m.groups() if x is not None]
                if len(g) >= 2:
                    prefix = norm_ref(g[0])
                    num = g[-1]
                    try:
                        num2 = str(int(str(num)))
                    except Exception:
                        num2 = str(num).lstrip("0") or str(num)
                    cand = prefix + num2
                else:
                    cand = norm_ref(m.group(0))

            if pos >= best_pos:
                best_pos = pos
                best = cand

    return best


# =========================
# Table detection (fast + reliable)
# =========================
HEADER_HINTS = {
    "date", "posting", "invoice", "inv", "document", "doc", "reference", "ref",
    "description", "details", "debit", "credit", "amount", "balance", "total", "type"
}


def looks_like_header(cell) -> bool:
    s = to_str(cell).lower()
    if not s:
        return False
    return any(k in s for k in HEADER_HINTS)


def row_non_empty_count(row: pd.Series) -> int:
    return int(row.notna().sum())


def detect_best_table_in_sheet(raw: pd.DataFrame, max_scan_rows: int = 80, sheet_name: str = ""):
    if raw is None or raw.empty:
        return None, None

    scan_rows = min(max_scan_rows, len(raw))
    best_score = -1
    best_df = None
    best_header_row = None

    for r in range(scan_rows):
        header_row = raw.iloc[r]
        non_empty = row_non_empty_count(header_row)
        if non_empty < 3:
            continue

        header_hits = sum(looks_like_header(x) for x in header_row.values)
        data_block = raw.iloc[r + 1: r + 1 + 30].copy()
        if data_block.empty:
            continue

        block_counts = [row_non_empty_count(data_block.iloc[i]) for i in range(min(len(data_block), 10))]
        consistency = float(np.mean(block_counts)) if block_counts else 0.0

        score = header_hits * 6 + non_empty + consistency
        if score > best_score:
            cols = [str(x).strip() if not pd.isna(x) else "" for x in header_row.values]
            df = raw.iloc[r + 1:].copy()
            df.columns = cols[: df.shape[1]]
            df = df.loc[:, [c for c in df.columns if str(c).strip() != ""]]
            df = df.dropna(how="all").reset_index(drop=True)
            best_score = score
            best_df = df
            best_header_row = r

    if best_df is None:
        return None, None

    meta = {
        "sheet_name": sheet_name,
        "header_row": int(best_header_row),
        "score": float(best_score),
        "rows": int(len(best_df)),
        "cols": int(len(best_df.columns)),
    }
    return best_df, meta


# =========================
# Column mapping (practical)
# =========================
SYNONYMS = {
    "txn_date": ["DATE", "POSTING DATE", "DOC DATE", "DOCUMENT DATE", "INVOICE DATE", "TRANSACTION DATE"],
    "invoice_no": ["INVOICE", "VENDOR INVOICE", "INVOICE NO", "INVOICE NUMBER", "SUPPLIER INVOICE", "INV NO", "INV#"],
    "doc_no": [
        "DOCUMENT", "DOCUMENT NO", "DOCUMENT NUMBER", "DOC NO", "DOC#", "VOUCHER", "VOUCHER NO",
        "EXTERNAL DOC", "EXTERNAL DOCUMENT", "EXTERNAL DOCUMENT NO", "EXTERNAL DOCUMENT NUMBER"
    ],
    "reference_text": ["REFERENCE", "REF", "REFERENCE NO", "OUR REF", "YOUR REF", "DOCUMENT NO", "DOC NO"],
    "description": ["DESCRIPTION", "DETAILS", "NARRATION", "TEXT", "PARTICULARS"],
    "debit": ["DEBIT", "DR", "DEBITS"],
    "credit": ["CREDIT", "CR", "CREDITS"],
    "amount": ["AMOUNT", "VALUE", "NET AMOUNT", "TOTAL", "AMOUNT (USD)", "AMOUNT (ZWG)", "AMT"],
    "balance": ["BALANCE", "RUNNING BALANCE"],
    "doc_type": ["TYPE", "DOCUMENT TYPE", "DOC TYPE"],
}


def infer_col_by_type(df: pd.DataFrame, role: str) -> str:
    best_col = ""
    best_score = -1e9

    for c in df.columns:
        s = df[c].map(to_str)
        n = df[c].map(to_num)
        header = to_str(c).lower()

        pct_num = float(n.notna().mean())
        avg_len = float(s.map(len).mean()) if len(s) else 0.0
        sparsity = float(df[c].isna().mean())

        score = 0.0
        if role == "date":
            sample = s.head(25).tolist()
            ok = 0
            for v in sample:
                if not v:
                    continue
                try:
                    dt_parse(v, fuzzy=True)
                    ok += 1
                except Exception:
                    pass
            pct_date = ok / max(1, len(sample))
            score = pct_date * 10 - sparsity * 2
            if "date" in header or "posting" in header:
                score += 2
        elif role == "amount":
            score = pct_num * 10 - sparsity * 1
            if "amount" in header or "amt" in header or "value" in header or "total" in header:
                score += 3
        elif role == "invoice_no":
            score = avg_len * 0.4 + (s.nunique(dropna=True) / max(1, len(s))) * 2
            if "invoice" in header or header.strip() in ("inv", "inv#"):
                score += 5
        elif role == "doc_no":
            score = avg_len * 0.5 + (s.nunique(dropna=True) / max(1, len(s))) * 1.5
            if "doc" in header or "external" in header or "reference" in header:
                score += 4
        elif role == "description":
            score = avg_len * 0.8
            if "desc" in header or "details" in header or "narr" in header or "particular" in header:
                score += 3

        if score > best_score:
            best_score = score
            best_col = c

    return best_col


def map_columns(df: pd.DataFrame) -> Dict[str, str]:
    cols = df.columns.tolist()
    col_norm = {c: upper(norm_space(c)) for c in cols}
    mapping: Dict[str, str] = {}

    for target, syns in SYNONYMS.items():
        found = ""
        for c, cn in col_norm.items():
            if cn in syns:
                found = c
                break
        if not found:
            for c, cn in col_norm.items():
                if any(s in cn for s in syns):
                    found = c
                    break

        if not found:
            if target in ("txn_date",):
                found = infer_col_by_type(df, "date")
            elif target in ("amount",):
                found = infer_col_by_type(df, "amount")
            elif target in ("invoice_no",):
                found = infer_col_by_type(df, "invoice_no")
            elif target in ("doc_no", "reference_text"):
                found = infer_col_by_type(df, "doc_no")
            elif target in ("description",):
                found = infer_col_by_type(df, "description")

        if found and found in df.columns:
            mapping[target] = found

    return mapping


# =========================
# Normalization (no currency gate)
# =========================
def normalize_table(
    df_raw: pd.DataFrame,
    meta: Dict[str, Any],
    colmap: Dict[str, str],
    source_type: str,
    source_file: str,
    cfg: ReconConfig,
) -> pd.DataFrame:
    df = df_raw.copy()

    txn_date = safe_col(df, colmap.get("txn_date", ""))
    invoice_no = safe_col(df, colmap.get("invoice_no", ""))
    doc_no = safe_col(df, colmap.get("doc_no", ""))
    reference_text = safe_col(df, colmap.get("reference_text", ""))
    description = safe_col(df, colmap.get("description", ""))
    doc_type = safe_col(df, colmap.get("doc_type", ""))

    debit = safe_col(df, colmap.get("debit", ""))
    credit = safe_col(df, colmap.get("credit", ""))
    amount = safe_col(df, colmap.get("amount", ""))
    balance = safe_col(df, colmap.get("balance", ""))

    out = pd.DataFrame()
    out["txn_date"] = txn_date.apply(lambda x: to_date(x, dayfirst=True))
    out["invoice_no_raw"] = invoice_no.apply(to_str)
    out["doc_no_raw"] = doc_no.apply(to_str)
    out["reference_text_raw"] = reference_text.apply(to_str)
    out["description_raw"] = description.apply(to_str)
    out["doc_type"] = doc_type.apply(to_str)

    out["invoice_key"] = out["invoice_no_raw"].apply(norm_ref)
    out["doc_key"] = out["doc_no_raw"].apply(norm_ref)

    d = debit.apply(to_num)
    c = credit.apply(to_num)
    a = amount.apply(to_num)

    if d.notna().any() or c.notna().any():
        out["amount_signed"] = d.fillna(0.0) - c.fillna(0.0)
    else:
        out["amount_signed"] = a

    if source_type == "ledger" and cfg.flip_ledger_sign:
        out["amount_signed"] = -out["amount_signed"]

    out["amt_r2"] = out["amount_signed"].apply(round2)
    out["amt_r0"] = out["amount_signed"].apply(round0)
    out["abs_amount"] = out["amount_signed"].abs()

    blob = (
        out["invoice_no_raw"].astype(str)
        + " "
        + out["doc_no_raw"].astype(str)
        + " "
        + out["reference_text_raw"].astype(str)
        + " "
        + out["description_raw"].astype(str)
        + " "
        + out["doc_type"].astype(str)
    )
    out["blob_text"] = blob.apply(norm_space)
    out["docid"] = out["blob_text"].apply(extract_docid)

    out["token_blob"] = out["blob_text"].apply(lambda s: " ".join(extract_tokens(s, min_len=cfg.min_token_len)))

    out["source_type"] = source_type
    out["source_file"] = source_file
    out["sheet_name"] = meta.get("sheet_name", "")
    out["table_score"] = float(meta.get("score", 0.0))

    out["row_uid"] = out.apply(
        lambda r: hash_row(
            source_file,
            meta.get("sheet_name", ""),
            r.get("txn_date", ""),
            r.get("invoice_key", ""),
            r.get("doc_key", ""),
            r.get("docid", ""),
            r.get("amt_r2", ""),
        ),
        axis=1,
    )

    keep = out["amount_signed"].notna() | out["invoice_key"].ne("") | out["doc_key"].ne("") | out["docid"].ne("")
    out = out.loc[keep].copy()

    out = out.loc[~(out["txn_date"].isna() & out["invoice_key"].eq("") & out["doc_key"].eq("") & out["docid"].eq(""))].copy()

    return out.reset_index(drop=True)


def process_workbook_bytes(xlsx_bytes: bytes, filename: str, source_type: str, cfg: ReconConfig) -> Tuple[pd.DataFrame, pd.DataFrame]:
    xls = pd.ExcelFile(io.BytesIO(xlsx_bytes), engine="openpyxl")
    all_norm = []
    logs = []

    for sh in xls.sheet_names:
        raw = pd.read_excel(io.BytesIO(xlsx_bytes), sheet_name=sh, header=None, dtype=object, engine="openpyxl")
        df, meta = detect_best_table_in_sheet(raw, sheet_name=sh)

        if df is None or df.empty or meta is None or meta["rows"] < 5 or meta["cols"] < 3:
            logs.append(
                {
                    "source_type": source_type,
                    "sheet_name": sh,
                    "status": "skipped",
                    "rows": 0,
                    "score": float(meta["score"]) if meta else 0.0,
                    "notes": "no table / too small",
                }
            )
            continue

        colmap = map_columns(df)
        norm = normalize_table(df, meta, colmap, source_type=source_type, source_file=filename, cfg=cfg)
        if norm.empty:
            logs.append(
                {
                    "source_type": source_type,
                    "sheet_name": sh,
                    "status": "skipped",
                    "rows": 0,
                    "score": float(meta["score"]),
                    "notes": "normalized empty",
                }
            )
            continue

        all_norm.append(norm)
        logs.append(
            {
                "source_type": source_type,
                "sheet_name": sh,
                "status": "ok",
                "rows": int(len(norm)),
                "score": float(meta["score"]),
                "notes": "",
            }
        )

    combined = pd.concat(all_norm, ignore_index=True) if all_norm else pd.DataFrame()
    return combined, pd.DataFrame(logs)


# =========================
# Matching (fast, workbook-friendly, no currency gate)
# =========================
def ensure_match_cols(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for c, d in [
        ("match_status", "unmatched"),
        ("match_group_id", None),
        ("match_rule", None),
        ("match_score", 0.0),
        ("match_reason", None),
        ("counterparty_uids", None),
    ]:
        if c not in out.columns:
            out[c] = d
    return out


def score_docid(amount_diff_abs: float, tol: float) -> Tuple[float, str]:
    if amount_diff_abs <= tol:
        return 0.98, "DocID totals match within tolerance"
    if amount_diff_abs <= tol * 5:
        return 0.75, "DocID matches but totals differ a bit"
    return 0.60, "DocID matches but totals differ a lot"


def score_invoice(amount_diff_abs: float, tol: float, date_diff: int, win: int, overlap: float) -> Tuple[float, str]:
    score = 0.94
    reason = "Invoice totals matched"
    if amount_diff_abs > tol:
        score -= min(0.25, (amount_diff_abs / max(0.01, tol)) * 0.05)
        reason = "Invoice matched but totals differ"
    if date_diff > win:
        score -= 0.10
        reason = reason + " and date is far"
    if overlap < 0.10:
        score -= 0.05
    return max(0.40, float(score)), reason


def date_diff_days(d1, d2) -> int:
    if pd.isna(d1) or pd.isna(d2):
        return 999999
    return abs((pd.to_datetime(d1).date() - pd.to_datetime(d2).date()).days)


def reconcile_fast(statement_all: pd.DataFrame, ledger_all: pd.DataFrame, cfg: ReconConfig) -> Dict[str, Any]:
    st_df = ensure_match_cols(statement_all)
    lg_df = ensure_match_cols(ledger_all)

    st_df["match_status"] = "unmatched"
    lg_df["match_status"] = "unmatched"
    st_df["match_group_id"] = None
    lg_df["match_group_id"] = None

    details_rows = []
    gid_num = 1

    def new_gid(prefix="G"):
        nonlocal gid_num
        gid = f"{prefix}{gid_num}"
        gid_num += 1
        return gid

    used_ledger = set()

    # 1) DOCID TOTALS
    st_doc = st_df.loc[st_df["docid"].fillna("").astype(str).ne("")].copy()
    lg_doc = lg_df.loc[lg_df["docid"].fillna("").astype(str).ne("")].copy()

    if not st_doc.empty or not lg_doc.empty:
        st_g = (
            st_doc.groupby("docid", as_index=False)
            .agg(
                statement_date=("txn_date", "min"),
                statement_ref=("docid", "first"),
                statement_details=("description_raw", "first"),
                statement_amount=("abs_amount", "sum"),
                statement_uids=("row_uid", lambda s: ",".join(s.astype(str).tolist()[:300])),
            )
        )
        lg_g = (
            lg_doc.groupby("docid", as_index=False)
            .agg(
                ledger_date=("txn_date", "min"),
                ledger_ref=("docid", "first"),
                ledger_details=("description_raw", "first"),
                ledger_amount=("abs_amount", "sum"),
                ledger_uids=("row_uid", lambda s: ",".join(s.astype(str).tolist()[:300])),
            )
        )

        merged = pd.merge(st_g, lg_g, on="docid", how="outer", indicator=True)
        merged["diff"] = merged["statement_amount"].fillna(0.0) - merged["ledger_amount"].fillna(0.0)
        merged["abs_diff"] = merged["diff"].abs()

        for _, r in merged.iterrows():
            if r["_merge"] != "both":
                continue

            sc, rs = score_docid(float(r["abs_diff"]), cfg.amount_tolerance)
            status = "Matched" if (float(r["abs_diff"]) <= cfg.amount_tolerance and sc >= cfg.min_auto_confidence) else "Needs review"

            gid = new_gid("D")
            s_uids = [x for x in to_str(r.get("statement_uids", "")).split(",") if x.strip()]
            l_uids = [x for x in to_str(r.get("ledger_uids", "")).split(",") if x.strip()]

            st_df.loc[
                st_df["row_uid"].astype(str).isin(s_uids),
                ["match_status", "match_group_id", "match_rule", "match_score", "match_reason", "counterparty_uids"],
            ] = ["matched", gid, "docid_total", sc, rs, ",".join(l_uids[:80])]

            lg_df.loc[
                lg_df["row_uid"].astype(str).isin(l_uids),
                ["match_status", "match_group_id", "match_rule", "match_score", "match_reason", "counterparty_uids"],
            ] = ["matched", gid, "docid_total", sc, rs, ",".join(s_uids[:80])]

            for lu in l_uids:
                used_ledger.add(str(lu))

            details_rows.append(
                {
                    "match_group_id": gid,
                    "match_rule": "docid_total",
                    "match_status": status,
                    "match_score": sc,
                    "match_reason": rs,
                    "match_key": r.get("docid"),
                    "statement_date": r.get("statement_date"),
                    "statement_ref": r.get("docid"),
                    "statement_amount": r.get("statement_amount"),
                    "ledger_date": r.get("ledger_date"),
                    "ledger_ref": r.get("docid"),
                    "ledger_amount": r.get("ledger_amount"),
                    "difference_statement_minus_ledger": r.get("diff"),
                }
            )

    # 2) INVOICE TOTALS
    st_rem = st_df.loc[~st_df["match_status"].eq("matched")].copy()
    lg_rem = lg_df.loc[~lg_df["match_status"].eq("matched")].copy()

    st_inv = st_rem.loc[st_rem["invoice_key"].fillna("").astype(str).ne("")].copy()
    lg_inv = lg_rem.loc[lg_rem["invoice_key"].fillna("").astype(str).ne("")].copy()

    if not st_inv.empty or not lg_inv.empty:
        st_g = (
            st_inv.groupby("invoice_key", as_index=False)
            .agg(
                statement_date=("txn_date", "min"),
                statement_ref=("invoice_no_raw", "first"),
                statement_details=("description_raw", "first"),
                statement_amount=("amount_signed", "sum"),
                statement_uids=("row_uid", lambda s: ",".join(s.astype(str).tolist()[:200])),
            )
        )
        lg_g = (
            lg_inv.groupby("invoice_key", as_index=False)
            .agg(
                ledger_date=("txn_date", "min"),
                ledger_ref=("invoice_no_raw", "first"),
                ledger_details=("description_raw", "first"),
                ledger_amount=("amount_signed", "sum"),
                ledger_uids=("row_uid", lambda s: ",".join(s.astype(str).tolist()[:200])),
            )
        )

        merged = pd.merge(st_g, lg_g, on="invoice_key", how="inner")
        if not merged.empty:
            merged["diff"] = merged["statement_amount"] - merged["ledger_amount"]
            merged["abs_diff"] = merged["diff"].abs()
            merged["date_diff"] = merged.apply(lambda x: date_diff_days(x["statement_date"], x["ledger_date"]), axis=1)
            merged["overlap"] = merged.apply(lambda x: token_overlap(to_str(x["statement_details"]), to_str(x["ledger_details"]), min_len=cfg.min_token_len), axis=1)

            merged = merged.sort_values(["abs_diff", "date_diff"], ascending=[True, True])

            for _, r in merged.iterrows():
                s_uids = [x for x in to_str(r.get("statement_uids", "")).split(",") if x.strip()]
                l_uids = [x for x in to_str(r.get("ledger_uids", "")).split(",") if x.strip()]

                if any(st_df.loc[st_df["row_uid"].astype(str).isin(s_uids), "match_status"].eq("matched")):
                    continue
                if any(lu in used_ledger for lu in l_uids):
                    continue

                sc, rs = score_invoice(float(r["abs_diff"]), cfg.amount_tolerance, int(r["date_diff"]), cfg.date_window_days, float(r["overlap"]))
                status = "Matched" if (float(r["abs_diff"]) <= cfg.amount_tolerance and sc >= cfg.min_auto_confidence) else "Needs review"

                gid = new_gid("I")

                st_df.loc[
                    st_df["row_uid"].astype(str).isin(s_uids),
                    ["match_status", "match_group_id", "match_rule", "match_score", "match_reason", "counterparty_uids"],
                ] = ["matched", gid, "invoice_total", sc, rs, ",".join(l_uids[:80])]

                lg_df.loc[
                    lg_df["row_uid"].astype(str).isin(l_uids),
                    ["match_status", "match_group_id", "match_rule", "match_score", "match_reason", "counterparty_uids"],
                ] = ["matched", gid, "invoice_total", sc, rs, ",".join(s_uids[:80])]

                for lu in l_uids:
                    used_ledger.add(str(lu))

                details_rows.append(
                    {
                        "match_group_id": gid,
                        "match_rule": "invoice_total",
                        "match_status": status,
                        "match_score": sc,
                        "match_reason": rs,
                        "match_key": r.get("invoice_key"),
                        "statement_date": r.get("statement_date"),
                        "statement_ref": r.get("statement_ref"),
                        "statement_amount": r.get("statement_amount"),
                        "ledger_date": r.get("ledger_date"),
                        "ledger_ref": r.get("ledger_ref"),
                        "ledger_amount": r.get("ledger_amount"),
                        "difference_statement_minus_ledger": r.get("diff"),
                    }
                )

    # 3) AMOUNT + DATE fallback
    st_rem = st_df.loc[~st_df["match_status"].eq("matched")].copy()
    lg_rem = lg_df.loc[~lg_df["match_status"].eq("matched")].copy()
    if not st_rem.empty and not lg_rem.empty:
        lg_rem = lg_rem.copy()
        lg_rem["bucket"] = lg_rem["amt_r0"].astype("float64")

        buckets = {}
        for b, g in lg_rem.groupby("bucket"):
            buckets[b] = g.head(cfg.max_candidates_per_bucket).copy()

        for idx, sr in st_rem.iterrows():
            s_amt0 = sr.get("amt_r0", np.nan)
            s_dt = sr.get("txn_date", pd.NaT)
            if pd.isna(s_amt0) or pd.isna(sr.get("amount_signed", np.nan)):
                continue

            cand = buckets.get(float(s_amt0))
            if cand is None or cand.empty:
                continue

            if pd.notna(s_dt):
                cand2 = cand.loc[cand["txn_date"].apply(lambda d: within_window(d, s_dt, cfg.date_window_days))].copy()
            else:
                cand2 = cand.copy()

            if cand2.empty:
                continue

            s_amt = float(sr.get("amount_signed"))
            cand2["abs_diff"] = (cand2["amount_signed"].astype(float) - s_amt).abs()
            cand2 = cand2.sort_values(["abs_diff"]).head(6)

            best = cand2.iloc[0]
            if float(best["abs_diff"]) > cfg.amount_tolerance:
                continue

            gid = new_gid("A")
            l_uid = str(best["row_uid"])

            if l_uid in used_ledger:
                continue

            overlap = token_overlap(to_str(sr.get("blob_text", "")), to_str(best.get("blob_text", "")), min_len=cfg.min_token_len)
            sc = 0.72 + (0.08 if overlap >= 0.20 else 0.0)
            rs = "Matched by amount + date window" + (" + tokens" if overlap >= 0.20 else "")

            st_df.loc[st_df.index == idx, ["match_status", "match_group_id", "match_rule", "match_score", "match_reason", "counterparty_uids"]] = [
                "matched", gid, "amount_date", sc, rs, l_uid
            ]
            lg_df.loc[lg_df["row_uid"].astype(str) == l_uid, ["match_status", "match_group_id", "match_rule", "match_score", "match_reason", "counterparty_uids"]] = [
                "matched", gid, "amount_date", sc, rs, str(sr["row_uid"])
            ]
            used_ledger.add(l_uid)

            details_rows.append(
                {
                    "match_group_id": gid,
                    "match_rule": "amount_date",
                    "match_status": "Matched" if sc >= cfg.min_auto_confidence else "Needs review",
                    "match_score": sc,
                    "match_reason": rs,
                    "match_key": f"AMT0={s_amt0}",
                    "statement_date": sr.get("txn_date"),
                    "statement_ref": sr.get("invoice_no_raw") or sr.get("doc_no_raw") or sr.get("docid"),
                    "statement_amount": sr.get("amount_signed"),
                    "ledger_date": best.get("txn_date"),
                    "ledger_ref": best.get("invoice_no_raw") or best.get("doc_no_raw") or best.get("docid"),
                    "ledger_amount": best.get("amount_signed"),
                    "difference_statement_minus_ledger": float(sr.get("amount_signed")) - float(best.get("amount_signed")),
                }
            )

    match_detail = pd.DataFrame(details_rows)

    matched_statement = st_df.loc[st_df["match_status"].eq("matched")].copy()
    unmatched_statement = st_df.loc[~st_df["match_status"].eq("matched")].copy()

    matched_ledger = lg_df.loc[lg_df["match_status"].eq("matched")].copy()
    unmatched_ledger = lg_df.loc[~lg_df["match_status"].eq("matched")].copy()

    left_table = pd.DataFrame({
        "Date": unmatched_ledger["txn_date"],
        "Ref": unmatched_ledger["docid"].replace("", np.nan)
            .combine_first(unmatched_ledger["invoice_no_raw"].replace("", np.nan))
            .combine_first(unmatched_ledger["doc_no_raw"].replace("", np.nan)),
        "Details": unmatched_ledger["description_raw"].fillna("").astype(str),
        "Amount": unmatched_ledger["amount_signed"],
        "Action": "Include on statement",
    })

    right_table = pd.DataFrame({
        "Date": unmatched_statement["txn_date"],
        "Ref": unmatched_statement["docid"].replace("", np.nan)
            .combine_first(unmatched_statement["invoice_no_raw"].replace("", np.nan))
            .combine_first(unmatched_statement["doc_no_raw"].replace("", np.nan)),
        "Details": unmatched_statement["description_raw"].fillna("").astype(str),
        "Amount": unmatched_statement["amount_signed"],
        "Action": "Post in ledger",
    })

    left_table = left_table.sort_values(["Date"], na_position="last").reset_index(drop=True)
    right_table = right_table.sort_values(["Date"], na_position="last").reset_index(drop=True)

    return {
        "statement_all": st_df,
        "ledger_all": lg_df,
        "matched_statement": matched_statement,
        "unmatched_statement": unmatched_statement,
        "matched_ledger": matched_ledger,
        "unmatched_ledger": unmatched_ledger,
        "match_detail": match_detail,
        "left_table": left_table,
        "right_table": right_table,
    }


# =========================
# Excel output (your Recon template structure + commentary)
# =========================
def excel_safe(v):
    if v is None:
        return None
    try:
        if pd.isna(v):
            return None
    except Exception:
        pass
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, np.datetime64):
        dtv = pd.to_datetime(v, errors="coerce")
        if pd.isna(dtv):
            return None
        return dtv.to_pydatetime()
    if isinstance(v, date) and not isinstance(v, datetime):
        return datetime(v.year, v.month, v.day)
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return None if np.isnan(v) else float(v)
    return v


def export_recon_template_bytes(
    template_bytes: bytes,
    supplier_name: str,
    as_at_dt: datetime,
    left_df: pd.DataFrame,
    right_df: pd.DataFrame,
    match_detail: pd.DataFrame,
    statement_all: pd.DataFrame,
    ledger_all: pd.DataFrame,
    commentary_lines: List[Tuple[str, str]],
):
    wb = load_workbook(io.BytesIO(template_bytes))

    ws = wb["Recon"] if "Recon" in wb.sheetnames else wb[wb.sheetnames[0]]

    try:
        ws["F2"].value = supplier_name
    except Exception:
        pass
    try:
        ws["K5"].value = as_at_dt
        ws["K5"].number_format = "dd/mm/yyyy"
    except Exception:
        pass

    start_row = 9
    totals_row = 29

    needed_rows = max(int(len(left_df)), int(len(right_df)), 1)
    available_rows = totals_row - start_row
    insert_n = max(0, needed_rows - available_rows)

    if insert_n > 0:
        ws.insert_rows(totals_row, amount=insert_n)
        totals_row += insert_n

    template_style_row = start_row
    for r in range(start_row + available_rows, start_row + needed_rows):
        for c in range(2, 13):
            src = ws.cell(template_style_row, c)
            dst = ws.cell(r, c)
            dst._style = pycopy(src._style)
            dst.number_format = src.number_format

    for r in range(start_row, start_row + needed_rows):
        for c in range(2, 13):
            ws.cell(r, c).value = None

    LEFT = {"Date": 2, "Ref": 3, "Details": 4, "Amount": 5, "Action": 6}
    RIGHT = {"Date": 8, "Ref": 9, "Details": 10, "Amount": 11, "Action": 12}
    date_fmt = "dd/mm/yy"

    left_df2 = left_df.copy() if left_df is not None else pd.DataFrame()
    right_df2 = right_df.copy() if right_df is not None else pd.DataFrame()

    for i in range(needed_rows):
        r = start_row + i
        if i < len(left_df2):
            it = left_df2.iloc[i].to_dict()
            ws.cell(r, LEFT["Date"]).value = excel_safe(it.get("Date"))
            ws.cell(r, LEFT["Date"]).number_format = date_fmt
            ws.cell(r, LEFT["Ref"]).value = excel_safe(it.get("Ref"))
            ws.cell(r, LEFT["Details"]).value = excel_safe(it.get("Details"))
            ws.cell(r, LEFT["Amount"]).value = excel_safe(it.get("Amount"))
            ws.cell(r, LEFT["Action"]).value = excel_safe(it.get("Action", ""))

        if i < len(right_df2):
            it = right_df2.iloc[i].to_dict()
            ws.cell(r, RIGHT["Date"]).value = excel_safe(it.get("Date"))
            ws.cell(r, RIGHT["Date"]).number_format = date_fmt
            ws.cell(r, RIGHT["Ref"]).value = excel_safe(it.get("Ref"))
            ws.cell(r, RIGHT["Details"]).value = excel_safe(it.get("Details"))
            ws.cell(r, RIGHT["Amount"]).value = excel_safe(it.get("Amount"))
            ws.cell(r, RIGHT["Action"]).value = excel_safe(it.get("Action", ""))

    left_sum_range = f"E{start_row}:E{start_row + needed_rows - 1}"
    right_sum_range = f"K{start_row}:K{start_row + needed_rows - 1}"
    ws[f"E{totals_row}"].value = f"=SUM({left_sum_range})"
    ws[f"K{totals_row}"].value = f"=SUM({right_sum_range})"

    if "Commentary" in wb.sheetnames:
        wb.remove(wb["Commentary"])
    ws_c = wb.create_sheet("Commentary")
    ws_c["A1"].value = "Commentary"

    r0 = 3
    for k, v in commentary_lines:
        ws_c.cell(r0, 1).value = k
        ws_c.cell(r0, 2).value = v
        r0 += 1
    ws_c.column_dimensions["A"].width = 34
    ws_c.column_dimensions["B"].width = 80

    def write_df_sheet(name: str, df: pd.DataFrame):
        if name in wb.sheetnames:
            wb.remove(wb[name])
        wsx = wb.create_sheet(name[:31])
        if df is None:
            df = pd.DataFrame()
        df2 = df.copy()
        df2 = df2.where(pd.notna(df2), None)
        wsx.append(list(df2.columns))
        for row in df2.itertuples(index=False):
            wsx.append([excel_safe(v) for v in list(row)])
        for c in range(1, len(df2.columns) + 1):
            wsx.column_dimensions[get_column_letter(c)].width = 18
        wsx.freeze_panes = "A2"
        wsx.auto_filter.ref = wsx.dimensions

    write_df_sheet("Match_Detail", match_detail)
    write_df_sheet("Statement_All", statement_all)
    write_df_sheet("Ledger_All", ledger_all)

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


# =========================
# Ollama chat
# =========================
def ollama_chat(messages: List[Dict[str, str]], model: str, host: str, timeout_s: int = 120) -> str:
    if requests is None:
        raise RuntimeError("requests is not installed. Install it with: pip install requests")
    payload = {"model": model, "messages": messages, "stream": False}
    r = requests.post(f"{host.rstrip('/')}/api/chat", json=payload, timeout=timeout_s)
    r.raise_for_status()
    data = r.json()
    return data.get("message", {}).get("content", "")


def ollama_is_up(host: str) -> Tuple[bool, str]:
    if requests is None:
        return False, "requests not installed"
    try:
        r = requests.get(f"{host.rstrip('/')}/api/tags", timeout=3)
        if r.status_code == 200:
            return True, ""
        return False, f"HTTP {r.status_code}"
    except Exception as e:
        return False, str(e)


def build_run_context(rr: Dict[str, Any]) -> str:
    if not rr or rr.get("error"):
        return ""
    ctx = {
        "counts": {
            "statement_lines": int(len(rr.get("statement_all", []))),
            "ledger_lines": int(len(rr.get("ledger_all", []))),
            "matched_statement": int(len(rr.get("matched_statement", []))),
            "unmatched_statement": int(len(rr.get("unmatched_statement", []))),
            "matched_ledger": int(len(rr.get("matched_ledger", []))),
            "unmatched_ledger": int(len(rr.get("unmatched_ledger", []))),
        }
    }
    us = rr.get("unmatched_statement")
    if isinstance(us, pd.DataFrame) and not us.empty:
        cols = ["txn_date", "invoice_no_raw", "doc_no_raw", "docid", "amount_signed", "description_raw"]
        cols = [c for c in cols if c in us.columns]
        sample = us.head(10)[cols].copy()
        if "txn_date" in sample.columns:
            sample["txn_date"] = sample["txn_date"].astype(str)
        ctx["unmatched_statement_sample"] = sample.to_dict(orient="records")
    return json.dumps(ctx, ensure_ascii=False)


# =========================
# UI
# =========================
st.set_page_config(page_title="Tarisai", layout="wide")

# Hard lock light UI and readable text everywhere (local + Streamlit Cloud)
st.markdown(
    f"""
<style>
:root {{
  --wf-red: {WF_RED};
  --wf-bg: {WF_BG};
  --wf-card: {WF_CARD};
  --wf-text: {WF_TEXT};
  --wf-muted: {WF_MUTED};
  --wf-border: {WF_BORDER};
}}

html, body {{
  background: var(--wf-bg) !important;
  color: var(--wf-text) !important;
}}

.stApp {{
  background: var(--wf-bg) !important;
  color: var(--wf-text) !important;
}}

.block-container {{
  padding-top: 2.4rem !important;
  padding-bottom: 2rem !important;
  max-width: 1250px !important;
  background: var(--wf-bg) !important;
}}

section[data-testid="stSidebar"] {{
  background: var(--wf-bg) !important;
}}
section[data-testid="stSidebar"] * {{
  color: var(--wf-text) !important;
}}

header[data-testid="stHeader"] {{
  background: var(--wf-bg) !important;
}}
footer {{
  background: var(--wf-bg) !important;
}}

#MainMenu, footer {{
  visibility: hidden;
}}

h1, h2, h3, h4, h5, h6,
p, span, label, div {{
  color: var(--wf-text) !important;
}}

[data-testid="stMarkdownContainer"] * {{
  color: var(--wf-text) !important;
}}

[data-testid="stCaptionContainer"] * {{
  color: var(--wf-muted) !important;
}}

.tarisai-hero,
.tarisai-card {{
  background: var(--wf-card) !important;
  border: 1px solid var(--wf-border) !important;
  border-radius: 16px !important;
}}

.tarisai-hero {{
  border-left: 6px solid var(--wf-red) !important;
}}

.tarisai-title {{
  font-size: 22px !important;
  font-weight: 800 !important;
  margin: 0 !important;
}}
.tarisai-sub {{
  font-size: 14px !important;
  color: var(--wf-muted) !important;
  margin-top: 6px !important;
  margin-bottom: 0 !important;
}}

[data-testid="stTabs"] {{
  background: var(--wf-bg) !important;
}}
[data-testid="stTabs"] * {{
  color: var(--wf-text) !important;
}}
[data-testid="stTabs"] button {{
  background: var(--wf-card) !important;
  border: 1px solid var(--wf-border) !important;
}}
[data-testid="stTabs"] button[aria-selected="true"] {{
  border-bottom: 3px solid var(--wf-red) !important;
}}

[data-testid="stExpander"] {{
  background: var(--wf-card) !important;
  border: 1px solid var(--wf-border) !important;
  border-radius: 14px !important;
}}
[data-testid="stExpander"] * {{
  color: var(--wf-text) !important;
}}

[data-testid="stFileUploader"] {{
  background: var(--wf-card) !important;
  border: 1px solid var(--wf-border) !important;
  border-radius: 14px !important;
  padding: 10px !important;
}}
[data-testid="stFileUploader"] * {{
  color: var(--wf-text) !important;
}}

[data-testid="stTextInput"],
[data-testid="stNumberInput"],
[data-testid="stSelectbox"],
[data-testid="stMultiselect"],
[data-testid="stDateInput"] {{
  background: var(--wf-bg) !important;
}}
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input {{
  background: white !important;
  color: var(--wf-text) !important;
  border: 1px solid var(--wf-border) !important;
}}
[data-testid="stSelectbox"] div,
[data-testid="stMultiselect"] div {{
  background: white !important;
  color: var(--wf-text) !important;
  border-color: var(--wf-border) !important;
}}

[data-testid="stWidgetLabel"] > div {{
  color: var(--wf-text) !important;
  font-weight: 700 !important;
}}

div.stButton > button {{
  background: var(--wf-red) !important;
  color: white !important;
  border: 1px solid var(--wf-red) !important;
  border-radius: 12px !important;
  font-weight: 800 !important;
}}
div.stButton > button:hover {{
  opacity: 0.92 !important;
}}

[data-testid="stDataFrame"] {{
  background: var(--wf-card) !important;
  border: 1px solid var(--wf-border) !important;
  border-radius: 14px !important;
  padding: 6px !important;
}}
[data-testid="stDataFrame"] * {{
  color: var(--wf-text) !important;
}}
[data-testid="stDataFrame"] div[role="grid"] {{
  background: white !important;
}}
[data-testid="stDataFrame"] div[role="columnheader"] {{
  background: var(--wf-bg) !important;
  color: var(--wf-text) !important;
}}
[data-testid="stDataFrame"] div[role="row"] {{
  background: white !important;
}}

[data-testid="stChatMessage"] {{
  background: var(--wf-bg) !important;
}}
[data-testid="stChatMessage"] > div {{
  background: var(--wf-card) !important;
  border: 1px solid var(--wf-border) !important;
  border-radius: 14px !important;
  padding: 10px !important;
}}
[data-testid="stChatMessage"] * {{
  color: var(--wf-text) !important;
}}

</style>
""",
    unsafe_allow_html=True,
)

st.markdown(
    """
<div class="tarisai-hero">
  <p class="tarisai-title">Tarisai</p>
  <p class="tarisai-sub">Reconciliation + local chat (Ollama). Upload workbooks, run matching, review exceptions, download your Recon template output with commentary.</p>
</div>
""",
    unsafe_allow_html=True,
)

if "recon_result" not in st.session_state:
    st.session_state.recon_result = None
if "template_bytes" not in st.session_state:
    st.session_state.template_bytes = None
if "chat_messages" not in st.session_state:
    st.session_state.chat_messages = []
if "last_cfg" not in st.session_state:
    st.session_state.last_cfg = ReconConfig(
        date_window_days=14,
        amount_tolerance=0.05,
        min_auto_confidence=0.70,
        flip_ledger_sign=True,
    )

tabs = st.tabs(["Run", "Review", "Chat", "Download"])

# =========================
# RUN TAB
# =========================
with tabs[0]:
    st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
    c1, c2 = st.columns([1, 1])

    with c1:
        st.subheader("Supplier")
        supplier_name = st.text_input("Supplier name", value="SUPPLIER")

    with c2:
        st.subheader("Recon template")
        template_file = st.file_uploader("Upload your Recon template (xlsx)", type=["xlsx"], key="templ_up")
        if template_file is not None:
            st.session_state.template_bytes = template_file.getvalue()
        st.caption("Use your Recon template to get the exact output layout you want.")

    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
    a, b = st.columns(2)
    with a:
        st.subheader("Supplier workbook")
        supplier_file = st.file_uploader("Upload supplier Excel", type=["xlsx", "xls"], key="supp_up")
    with b:
        st.subheader("Ledger workbook")
        ledger_file = st.file_uploader("Upload ledger Excel", type=["xlsx", "xls"], key="ledg_up")
    st.markdown("</div>", unsafe_allow_html=True)

    st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
    st.subheader("Matching settings")

    cfg0 = st.session_state.last_cfg
    flip_sign = st.checkbox("Flip ledger sign (common when ledger sign is opposite)", value=bool(cfg0.flip_ledger_sign))
    date_window = st.number_input("Date window (days)", min_value=0, max_value=90, value=int(cfg0.date_window_days), step=1)
    tol = st.number_input("Amount tolerance", min_value=0.0, value=float(cfg0.amount_tolerance), step=0.01)
    min_conf = st.slider("Auto-match confidence", min_value=0.50, max_value=0.99, value=float(cfg0.min_auto_confidence), step=0.01)

    cfg = ReconConfig(
        date_window_days=int(date_window),
        amount_tolerance=float(tol),
        min_auto_confidence=float(min_conf),
        flip_ledger_sign=bool(flip_sign),
    )
    st.session_state.last_cfg = cfg

    run_clicked = st.button("Run reconciliation", use_container_width=True, disabled=not (supplier_file and ledger_file))
    st.markdown("</div>", unsafe_allow_html=True)

    if run_clicked:
        with st.spinner("Scanning workbooks and matching..."):
            s_bytes = supplier_file.getvalue()
            l_bytes = ledger_file.getvalue()

            st_raw, log_s = process_workbook_bytes(s_bytes, supplier_file.name, "statement", cfg)
            lg_raw, log_l = process_workbook_bytes(l_bytes, ledger_file.name, "ledger", cfg)
            run_log = pd.concat([log_s, log_l], ignore_index=True)

            if st_raw.empty or lg_raw.empty:
                st.session_state.recon_result = {
                    "error": "Extraction failed for one of the files. Check your sheets and headers.",
                    "run_log": run_log,
                }
            else:
                rr = reconcile_fast(st_raw, lg_raw, cfg)
                rr["error"] = None
                rr["run_log"] = run_log
                rr["supplier_name"] = supplier_name
                rr["as_at"] = datetime.now()
                st.session_state.recon_result = rr

        rr = st.session_state.recon_result
        if rr and not rr.get("error"):
            st.success("Done. Go to Review.")
        elif rr and rr.get("error"):
            st.error(rr["error"])

# =========================
# REVIEW TAB
# =========================
with tabs[1]:
    st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
    st.subheader("Review")

    rr = st.session_state.recon_result
    if not rr or rr.get("error"):
        st.info("Run a reconciliation first.")
        st.markdown("</div>", unsafe_allow_html=True)
    else:
        st_all = rr["statement_all"]
        lg_all = rr["ledger_all"]

        m1, m2, m3, m4 = st.columns(4)
        m1.metric("Statement lines", int(len(st_all)))
        m2.metric("Ledger lines", int(len(lg_all)))
        m3.metric("Matched", int(len(rr["matched_statement"])))
        m4.metric("Unmatched", int(len(rr["unmatched_statement"])))

        st.write("")
        st.subheader("Recon tables (this feeds your template output)")
        c1, c2 = st.columns(2)
        with c1:
            st.write("Ledger items missing on statement")
            st.dataframe(rr["left_table"].head(700), use_container_width=True, hide_index=True)
        with c2:
            st.write("Statement items missing in ledger")
            st.dataframe(rr["right_table"].head(700), use_container_width=True, hide_index=True)

        st.write("")
        with st.expander("Match detail (top 3000)", expanded=False):
            md = rr.get("match_detail", pd.DataFrame())
            if isinstance(md, pd.DataFrame) and not md.empty:
                st.dataframe(md.sort_values(["match_score"], ascending=False).head(3000), use_container_width=True, hide_index=True)
            else:
                st.info("No match detail created yet.")

        st.write("")
        with st.expander("Run log (sheet detection)", expanded=False):
            st.dataframe(rr.get("run_log", pd.DataFrame()), use_container_width=True, hide_index=True)

    st.markdown("</div>", unsafe_allow_html=True)

# =========================
# CHAT TAB (Ollama)
# =========================
with tabs[2]:
    st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
    st.subheader("Chat (Ollama local)")

    rr = st.session_state.recon_result
    has_ctx = rr is not None and not rr.get("error")

    c1, c2 = st.columns([1, 1])
    with c1:
        ollama_model = st.text_input("Model", value="llama3.2:3b")
    with c2:
        ollama_host = st.text_input("Ollama host", value="http://localhost:11434")

    is_up, up_err = ollama_is_up(ollama_host)
    with st.expander("Ollama debug", expanded=False):
        st.write("requests installed:", requests is not None)
        st.write("Host reachable:", is_up)
        if up_err:
            st.write("Detail:", up_err)

    if not is_up:
        st.info("Start Ollama, then come back here. Reconciliation still works without chat.")

    if has_ctx:
        st.caption("Ask why items are unmatched, what to fix, or what to do next.")
    else:
        st.caption("Run reconciliation first if you want the chat to explain your results.")

    for msg in st.session_state.chat_messages:
        with safe_chat_message(msg["role"]):
            st.write(msg["content"])

    user_prompt = safe_chat_input("Ask something about your reconciliation", key="tarisai_chat")
    if user_prompt:
        st.session_state.chat_messages.append({"role": "user", "content": user_prompt})
        with safe_chat_message("user"):
            st.write(user_prompt)

        if not is_up:
            bot_text = "Ollama is not reachable. Start it and confirm the host is http://localhost:11434"
            st.session_state.chat_messages.append({"role": "assistant", "content": bot_text})
            with safe_chat_message("assistant"):
                st.write(bot_text)
        else:
            ctx = build_run_context(rr) if has_ctx else ""
            sys_msg = (
                "You are Tarisai, a reconciliation assistant. "
                "Explain results like a finance analyst. "
                "Give clear actions: what to post, what to request from supplier, what to investigate in the ERP. "
                "If the user asks why unmatched, use the run context JSON."
            )

            messages = [{"role": "system", "content": sys_msg}]
            if ctx:
                messages.append({"role": "system", "content": f"Run context JSON: {ctx}"})
            for m in st.session_state.chat_messages[-12:]:
                messages.append({"role": m["role"], "content": m["content"]})

            with safe_chat_message("assistant"):
                try:
                    bot_text = ollama_chat(messages, model=ollama_model, host=ollama_host, timeout_s=180)
                except Exception as e:
                    bot_text = f"Ollama chat failed: {e}"

                st.write(bot_text)
                st.session_state.chat_messages.append({"role": "assistant", "content": bot_text})

    st.markdown("</div>", unsafe_allow_html=True)

# =========================
# DOWNLOAD TAB
# =========================
with tabs[3]:
    st.markdown('<div class="tarisai-card">', unsafe_allow_html=True)
    st.subheader("Download")

    rr = st.session_state.recon_result
    if not rr or rr.get("error"):
        st.info("Run a reconciliation first.")
        st.markdown("</div>", unsafe_allow_html=True)
    else:
        template_bytes = st.session_state.template_bytes
        if not template_bytes:
            st.warning("Upload your Recon template in the Run tab to get your exact output structure.")
            st.caption("You can still export raw data by using your template.")
        else:
            cfg = st.session_state.last_cfg
            commentary = [
                ("What matched", f"{len(rr['matched_statement'])} statement lines matched using DocID totals, invoice totals, and amount/date fallback."),
                ("What is outstanding", f"{len(rr['unmatched_statement'])} statement lines and {len(rr['unmatched_ledger'])} ledger lines still unmatched."),
                ("What to do next", "Check the Recon sheet items. Post missing invoices, request missing statements, and review amount mismatches."),
                ("Settings used", f"date window={cfg.date_window_days} days, tolerance={cfg.amount_tolerance}, min confidence={cfg.min_auto_confidence}, flip ledger sign={cfg.flip_ledger_sign}"),
                ("DocID handling", "Tarisai extracts DocIDs from invoice/doc/reference/description using multiple patterns, so it still matches when formats change."),
            ]

            try:
                out_bytes = export_recon_template_bytes(
                    template_bytes=template_bytes,
                    supplier_name=rr.get("supplier_name", "SUPPLIER"),
                    as_at_dt=rr.get("as_at", datetime.now()),
                    left_df=rr.get("left_table", pd.DataFrame()),
                    right_df=rr.get("right_table", pd.DataFrame()),
                    match_detail=rr.get("match_detail", pd.DataFrame()),
                    statement_all=rr.get("statement_all", pd.DataFrame()),
                    ledger_all=rr.get("ledger_all", pd.DataFrame()),
                    commentary_lines=commentary,
                )

                out_name = f"Tarisai_Recon_{rr.get('supplier_name','SUPPLIER')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                st.download_button(
                    "Download reconciliation workbook",
                    data=out_bytes,
                    file_name=out_name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.caption("Output includes: Recon (your template layout), Commentary, Match_Detail, Statement_All, Ledger_All.")
            except Exception as e:
                st.error(f"Failed to prepare Excel: {e}")

    st.markdown("</div>", unsafe_allow_html=True)
